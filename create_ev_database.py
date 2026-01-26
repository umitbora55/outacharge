from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Stiller
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='1F4E79')
turkey_fill = PatternFill('solid', fgColor='C00000')
category_fill = PatternFill('solid', fgColor='2E75B6')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill=header_fill):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = left_align if col <= 2 else center_align

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

# ===================== SAYFA 1: ÖZET =====================
ws = wb.active
ws.title = "ÖZET"

ws['A1'] = "ELEKTRİKLİ ARAÇ VERİTABANI KATALOGU"
ws['A1'].font = Font(bold=True, size=16, color='1F4E79')
ws['A2'] = "Tarih: Ocak 2026 | Toplam: 150+ Kaynak | 12 Kategori"
ws['A2'].font = Font(italic=True, size=11)

headers = ['Kategori', 'Kaynak Sayısı', 'Tahmini Araç', 'Veri Kalitesi', 'İndirme Kolaylığı', 'En İyi Kaynak', 'Türkiye Üretimi']
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 7)

data = [
    ['🚗 Binek Araçlar', '15+', '500-900', '⭐⭐⭐⭐⭐', '⭐⭐⭐⭐⭐', 'Open EV Data, EVKX', 'TOGG T10X, T10F'],
    ['🏍️ Motosikletler', '8+', '42,000+', '⭐⭐⭐⭐', '⭐⭐⭐', 'TeoAlida, Motorwatt', '-'],
    ['🛵 E-Scooter/Moped', '10+', '500+', '⭐⭐⭐', '⭐⭐', 'NIU, Yadea, Gogoro', '-'],
    ['🚲 E-Bisiklet', '7+', '50,000+', '⭐⭐⭐⭐', '⭐⭐', 'Bikes.fan', '-'],
    ['🚌 E-Otobüs', '8+', '200+', '⭐⭐⭐⭐', '⭐⭐', 'Sustainable Bus', 'Karsan e-ATAK, e-ATA'],
    ['🚚 E-Kamyon/Van', '10+', '100+', '⭐⭐⭐', '⭐⭐', 'BYD, Volvo', 'Ford E-Transit'],
    ['⛵ E-Tekne', '8+', '50+', '⭐⭐⭐', '⭐⭐', 'Plugboats, Candela', '-'],
    ['✈️ eVTOL', '12+', '100+', '⭐⭐⭐⭐⭐', '⭐⭐⭐⭐', 'eVTOL.news', '-'],
    ['🏗️ İş Makineleri', '10+', '50+', '⭐⭐⭐⭐', '⭐⭐⭐', 'HeavyEquipmentData', '-'],
    ['🏎️ ATV/UTV/Golf', '8+', '100+', '⭐⭐⭐', '⭐⭐', 'Polaris, Club Car', '-'],
    ['♿ Mobilite Cihazları', '8+', '200+', '⭐⭐⭐', '⭐⭐', 'WHILL, Pride', '-'],
    ['🔌 Şarj Verisi', '10+', '500K+ kayıt', '⭐⭐⭐⭐', '⭐⭐⭐', 'Figshare, Caltech', 'Vestel, Gersan, ZES'],
]

for i, row in enumerate(data, 5):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
style_data(ws, 5, 16, 7)
set_column_widths(ws, [22, 14, 14, 16, 18, 25, 25])

# ===================== SAYFA 2: BİNEK ARAÇLAR =====================
ws2 = wb.create_sheet("Binek Araçlar")
headers = ['Kaynak Adı', 'URL', 'Araç Sayısı', 'Format', 'Lisans', 'Özellikler', 'Türkiye Desteği']
for i, h in enumerate(headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, 7)

data = [
    ['Open EV Data', 'github.com/chargeprice/open-ev-data', '500+', 'JSON', 'MIT → Ücretli', 'Şarj eğrileri, batarya kapasitesi', 'Evet'],
    ['EVKX GitHub', 'github.com/evkx/evkx.github.io', '400+', 'Markdown/JSON', 'Açık', 'Detaylı şarj eğrileri, menzil, tüketim', 'Evet'],
    ['EV-Database.org', 'ev-database.org/data-services-api', '300+', 'JSON/API', 'Ticari', 'Kapsamlı specs, fiyatlar, karşılaştırma', 'Kısmi'],
    ['US Car Models Data', 'github.com/abhionlyone/us-car-models-data', '15,000+', 'CSV', 'MIT', '1992-2026 arası modeller', 'Hayır'],
    ['Open Vehicle DB', 'github.com/plowman/open-vehicle-db', '10,000+', 'SQL/JSON', 'MIT', 'Make/Model/Year/Style', 'Hayır'],
    ['Vehicle Make Model Data', 'github.com/arthurkao/vehicle-make-model-data', '19,722', 'MySQL/JSON/CSV', 'MIT', '2001-2015 arası, Motosiklet dahil', 'Hayır'],
    ['EPA Fuel Economy', 'fueleconomy.gov/feg/download.shtml', 'ABD tüm', 'CSV', 'Kamu', 'Resmi ABD verileri', 'Hayır'],
    ['EEA CO2 Database', 'eea.europa.eu/en/datahub', 'AB tüm', 'CSV/SQL', 'Kamu', 'AB resmi emisyon verileri', 'Hayır'],
    ['AFDC Vehicle Data', 'afdc.energy.gov/vehicles', 'ABD tüm', 'PDF/Excel', 'Kamu', 'Alternatif yakıt araçları', 'Hayır'],
    ['VehicleDatabases EV API', 'vehicledatabases.com/electric-vehicle-specifications-api', 'Kapsamlı', 'API', 'Ücretli', 'VIN decode, tam specs', 'Hayır'],
    ['API Ninjas EV', 'api-ninjas.com/api/electricvehicle', '1000+', 'API', 'Freemium', 'Basit sorgular', 'Hayır'],
    ['High Mobility', 'high-mobility.com', 'Çoklu marka', 'API', 'Ücretli', 'Gerçek zamanlı telemetri', 'Hayır'],
    ['Chargeprice API', 'chargeprice.com', '500+', 'API', 'Ücretli', 'Şarj fiyatları + araç specs', 'Evet'],
    ['CarNewsChina', 'data.carnewschina.com', '3,620+', 'Web', 'Ücretsiz', 'Çin EV pazarı', 'Hayır'],
    ['VAHAN (Hindistan)', 'vahan.parivahan.gov.in', 'Hindistan tüm', 'Web', 'Kamu', 'Hindistan kayıt verileri', 'Hayır'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val)
style_data(ws2, 2, len(data)+1, 7)
set_column_widths(ws2, [25, 45, 12, 15, 15, 40, 15])

# ===================== SAYFA 3: MOTOSİKLETLER =====================
ws3 = wb.create_sheet("Motosikletler")
headers = ['Kaynak Adı', 'URL', 'Araç Sayısı', 'Format', 'Lisans', 'Özellikler']
for i, h in enumerate(headers, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, 6)

data = [
    ['Motorwatt EV Database', 'ev.motorwatt.com/ev-database/catalog-electric-motorcycles', '140+', 'Web', 'Ücretsiz', 'Elektrikli motosiklet specs'],
    ['TeoAlida Motorcycle DB', 'teoalida.com/cardatabase/motorcycles', '42,565', 'Excel/CSV/SQL', 'Ücretli', '607 marka, 1894-2025'],
    ['Moto.Car2db', 'moto.car2db.com', '30,000+', 'MySQL/CSV', 'Ücretli', 'Kapsamlı specs'],
    ['Moto-Data.net', 'moto-data.net', '15,000+', 'Web', 'Ücretsiz', 'Teknik specs, fotoğraflar'],
    ['AutoEvolution Moto', 'autoevolution.com/moto', '20,000+', 'Web', 'Ücretsiz', 'Tam specs, fotoğraflar'],
    ['MotorcycleDB', 'motorcycledb.com', '25,000+', 'Web', 'Ücretsiz', 'Hız, yakıt, boyutlar'],
    ['MotorcyclesData', 'motorcyclesdata.com', 'Satış verisi', 'Web', 'Ücretsiz', 'Pazar analizi, satış'],
    ['Zero Motorcycles', 'zeromotorcycles.com', '10+', 'Web', 'Resmi', 'ABD elektrikli moto'],
    ['Energica', 'energicamotor.com', '5+', 'Web', 'Resmi', 'İtalya premium EV moto'],
    ['LiveWire (Harley)', 'livewire.com', '3+', 'Web', 'Resmi', 'Harley elektrikli'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws3.cell(row=i, column=j, value=val)
style_data(ws3, 2, len(data)+1, 6)
set_column_widths(ws3, [25, 50, 12, 15, 12, 35])

# ===================== SAYFA 4: E-SCOOTER/MOPED =====================
ws4 = wb.create_sheet("E-Scooter Moped")
headers = ['Marka', 'URL', 'Menşe', 'Model Sayısı', 'Pazar Payı', 'Öne Çıkan Modeller', 'Türkiye Satışı']
for i, h in enumerate(headers, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, 7)

data = [
    ['Yadea', 'yadea.com', 'Çin', '50+', '#1 Dünya', 'G5, C1S, VFV, VFD', 'Evet'],
    ['NIU', 'niu.com', 'Çin', '30+', '#2 Premium', 'NQi, MQi GT EVO, UQi, RQi Sport', 'Evet'],
    ['Gogoro', 'gogoro.com', 'Tayvan', '20+', '#1 Tayvan', 'JEGO, CrossOver', 'Hayır'],
    ['Ather Energy', 'atherenergy.com', 'Hindistan', '5+', '#1 Hindistan', '450X, 450 Apex', 'Hayır'],
    ['Ola Electric', 'olaelectric.com', 'Hindistan', '5+', 'Hindistan', 'S1 Pro, S1 Air', 'Hayır'],
    ['Vmoto/Super Soco', 'vmoto.com', 'Avustralya/Çin', '15+', 'Global', 'CPx, TC Max', 'Evet'],
    ['Hero Electric', 'heroelectric.in', 'Hindistan', '10+', 'Hindistan', 'Optima, Nyx', 'Hayır'],
    ['BMW CE', 'bmw-motorrad.com', 'Almanya', '2+', 'Premium', 'CE 04, CE 02', 'Evet'],
    ['Vespa Elettrica', 'vespa.com', 'İtalya', '2+', 'Premium', 'Elettrica', 'Evet'],
    ['Silence', 'silence.eco', 'İspanya', '5+', 'Avrupa', 'S01, S02', 'Hayır'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val)
style_data(ws4, 2, len(data)+1, 7)
set_column_widths(ws4, [20, 30, 15, 12, 15, 35, 15])

# ===================== SAYFA 5: E-BİSİKLET =====================
ws5 = wb.create_sheet("E-Bisiklet")
headers = ['Kaynak Adı', 'URL', 'Araç Sayısı', 'Format', 'Lisans', 'Özellikler']
for i, h in enumerate(headers, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, 6)

data = [
    ['Bikes.fan', 'bikes.fan', '50,000+', 'Web', 'Ücretsiz', 'En kapsamlı, bileşenler, geometri'],
    ['UsableDatabases Bicycles', 'usabledatabases.com/database/bicycles-in-types', '24,496', 'Excel/CSV', 'Ücretli', '320 marka, 24 kategori'],
    ['ElectricBikeReview', 'electricbikereview.com', '1000+', 'Web', 'Ücretsiz', 'Detaylı incelemeler'],
    ['MTB Database', 'mtbdatabase.com/e-bikes', 'E-MTB', 'Web', 'Ücretsiz', 'Dağ bisikleti odaklı'],
    ['Road Bike Database', 'roadbikedatabase.com/e-bikes', 'E-Road', 'Web', 'Ücretsiz', 'Yol bisikleti odaklı'],
    ['ebikecomparisondatabase', 'ebikecomparisondatabase.com', '50+', 'Web', 'Ücretsiz', 'Karşılaştırma'],
    ['Bike Index API', 'bikeindex.org/documentation/api_v3', 'Registry', 'REST API', 'Ücretsiz', 'Kayıp bisiklet registry'],
    ['Bosch eBike (Motor)', 'bosch-ebike.com', 'Motor', 'Web', 'Resmi', 'Performance Line CX, Active'],
    ['Shimano Steps (Motor)', 'productinfo.shimano.com', 'Motor', 'Web', 'Resmi', 'EP8, E6100, E5000'],
    ['Bafang (Motor)', 'bafang-e.com', 'Motor', 'Web', 'Resmi', 'Ultra M620, M500'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws5.cell(row=i, column=j, value=val)
style_data(ws5, 2, len(data)+1, 6)
set_column_widths(ws5, [25, 45, 12, 12, 12, 40])

# ===================== SAYFA 6: E-OTOBÜS =====================
ws6 = wb.create_sheet("E-Otobüs")
headers = ['Üretici', 'URL', 'Menşe', 'Teslimat', 'Öne Çıkan Modeller', 'Türkiye Üretimi']
for i, h in enumerate(headers, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, 6)

data = [
    ['KARSAN', 'karsan.com', 'TÜRKİYE 🇹🇷', '5,000+', 'e-Jest, e-ATAK, e-ATA', 'EVET - Bursa'],
    ['OTOKAR', 'otokar.com.tr', 'TÜRKİYE 🇹🇷', '1,000+', 'e-Kent, e-Navigo', 'EVET - Sakarya'],
    ['BMC', 'bmc.com.tr', 'TÜRKİYE 🇹🇷', '500+', 'Procity Electric', 'EVET - İzmir'],
    ['BYD', 'bydeurope.com/pdp-bus-coach', 'Çin', '100,000+', 'K7, K8, K9, K11', 'Hayır'],
    ['Volvo Buses', 'volvobuses.com', 'İsveç', '50,000+', '7900 Electric, BZL', 'Hayır'],
    ['Mercedes eCitaro', 'mercedes-benz-bus.com', 'Almanya', '1,000+', 'eCitaro, eCitaro G', 'Hayır'],
    ['Solaris', 'solarisbus.com', 'Polonya', '3,000+', 'Urbino Electric', 'Hayır'],
    ['VDL', 'vdlbuscoach.com', 'Hollanda', '1,500+', 'Citea Electric', 'Hayır'],
    ['Yutong', 'yutong.com', 'Çin', '150,000+', 'E12, E18', 'Hayır'],
    ['Proterra', 'proterra.com', 'ABD', '1,000+', 'Catalyst', 'Hayır'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        cell = ws6.cell(row=i, column=j, value=val)
        if 'TÜRKİYE' in str(val) or 'EVET' in str(val):
            cell.font = Font(bold=True, color='C00000')
style_data(ws6, 2, len(data)+1, 6)
set_column_widths(ws6, [20, 35, 15, 12, 35, 18])

# ===================== SAYFA 7: E-KAMYON =====================
ws7 = wb.create_sheet("E-Kamyon Van")
headers = ['Üretici', 'URL', 'Menşe', 'Modeller', 'Özellikler', 'Türkiye']
for i, h in enumerate(headers, 1):
    ws7.cell(row=1, column=i, value=h)
style_header(ws7, 1, 6)

data = [
    ['Ford Otosan', 'ford.com.tr', 'TÜRKİYE 🇹🇷', 'E-Transit, E-Tourneo', 'Türkiye üretimi, Avrupa ihracatı', 'EVET - Kocaeli'],
    ['BYD Trucks', 'en.byd.com/truck', 'Çin', '8TT, T10', '12,000+ global, 400+ kWh', 'Planlanan (1B USD yatırım)'],
    ['Volvo Trucks', 'volvotrucks.com', 'İsveç', 'FM, FMX, FE, FL Electric', 'VNR Electric (ABD)', 'Hayır'],
    ['Daimler/RIZON', 'rizontrucks.com', 'Almanya/ABD', 'RIZON', 'Orta sınıf elektrikli', 'Hayır'],
    ['Tesla Semi', 'tesla.com/semi', 'ABD', 'Semi', '500 mil menzil', 'Hayır'],
    ['Rivian', 'rivian.com', 'ABD', 'EDV (Amazon)', 'Teslimat van', 'Hayır'],
    ['Freightliner', 'freightliner.com', 'ABD', 'eCascadia', 'Ağır yük', 'Hayır'],
    ['Scania', 'scania.com', 'İsveç', 'Electric trucks', 'Avrupa pazarı', 'Hayır'],
    ['MAN', 'man.eu', 'Almanya', 'eTGX, eTGS', 'Avrupa pazarı', 'Hayır'],
    ['IVECO', 'iveco.com', 'İtalya', 'eDaily', 'Hafif ticari', 'Hayır'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        cell = ws7.cell(row=i, column=j, value=val)
        if 'TÜRKİYE' in str(val) or 'EVET' in str(val):
            cell.font = Font(bold=True, color='C00000')
style_data(ws7, 2, len(data)+1, 6)
set_column_widths(ws7, [18, 30, 15, 25, 35, 25])

# ===================== SAYFA 8: E-TEKNE =====================
ws8 = wb.create_sheet("E-Tekne")
headers = ['Üretici', 'URL', 'Menşe', 'Modeller', 'Özellikler']
for i, h in enumerate(headers, 1):
    ws8.cell(row=1, column=i, value=h)
style_header(ws8, 1, 5)

data = [
    ['Candela', 'candela.com', 'İsveç', 'C-7, C-8, P-12', 'Hydrofoil, 57 nm menzil, 22 knot'],
    ['X Shore', 'xshore.com', 'İsveç', 'Eelex 8000', 'Premium elektrikli'],
    ['Torqeedo', 'torqeedo.com', 'Almanya', 'Travel, Cruise, Deep Blue', 'Motor üreticisi, 120 ft'],
    ['Vision Marine', 'visionmarinetechnologies.com', 'Kanada', 'E-Motion', '116 mph rekor'],
    ['Pure Watercraft', 'purewatercraft.com', 'ABD', 'Pure Outboard', 'Dıştan takma motor'],
    ['Evoy', 'evoy.no', 'Norveç', 'Outboard motors', 'Yüksek güç'],
    ['Forza X1', 'forzax1.com', 'ABD', 'FX-1', 'Spor tekne'],
    ['Plugboats (Kaynak)', 'plugboats.com', 'Global', '50+ tekne', 'Veritabanı, Gussies Awards'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws8.cell(row=i, column=j, value=val)
style_data(ws8, 2, len(data)+1, 5)
set_column_widths(ws8, [20, 40, 12, 25, 40])

# ===================== SAYFA 9: eVTOL =====================
ws9 = wb.create_sheet("eVTOL Hava Araçları")
headers = ['Şirket', 'URL', 'Menşe', 'Model', 'Menzil', 'Hız', 'Kapasite', 'Durum']
for i, h in enumerate(headers, 1):
    ws9.cell(row=1, column=i, value=h)
style_header(ws9, 1, 8)

data = [
    ['Joby Aviation', 'jobyaviation.com', 'ABD', 'S4', '150 mil', '200 mph', '5 kişi', 'FAA sertifikasyon 2025'],
    ['Archer Aviation', 'archer.com', 'ABD', 'Midnight', '100 mil', '150 mph', '4+1 kişi', '2028 LA Olympics hedef'],
    ['Beta Technologies', 'beta.team', 'ABD', 'ALIA', '250 mil', '170 mph', 'Kargo+yolcu', 'Geliştirme'],
    ['Volocopter', 'volocopter.com', 'Almanya', 'VoloCity', '18 mil', '62 mph', '2 kişi', 'Varlıklar satıldı (2025)'],
    ['EHang', 'ehang.com', 'Çin', 'EH216', '21 mil', '80 mph', '2 kişi', 'CAAC sertifikalı'],
    ['Vertical Aerospace', 'vertical-aerospace.com', 'İngiltere', 'VX4', '100 mil', '150 mph', '4+1 kişi', 'Geliştirme'],
    ['Wisk Aero', 'wisk.aero', 'ABD', 'Cora', '25 mil', '100 mph', '2 kişi', 'Boeing ortaklığı, otonom'],
    ['Eve Air Mobility', 'eveairmobility.com', 'Brezilya', 'Eve', '60 mil', '150 mph', '4+1 kişi', '2,850 sipariş'],
    ['SkyDrive', 'skydrive2020.com', 'Japonya', 'SD-05', '10 km', '100 km/h', '2 kişi', 'Osaka Expo 2025'],
    ['eVTOL News (Kaynak)', 'evtol.news', 'Global', '100+ proje', '-', '-', '-', 'Veritabanı'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws9.cell(row=i, column=j, value=val)
style_data(ws9, 2, len(data)+1, 8)
set_column_widths(ws9, [20, 30, 12, 12, 10, 10, 12, 25])

# ===================== SAYFA 10: İŞ MAKİNELERİ =====================
ws10 = wb.create_sheet("İş Makineleri")
headers = ['Üretici', 'URL', 'Elektrikli Modeller', 'Batarya', 'Özellikler']
for i, h in enumerate(headers, 1):
    ws10.cell(row=1, column=i, value=h)
style_header(ws10, 1, 5)

data = [
    ['Caterpillar', 'cat.com', '320 Electric', '387 kWh', '8 saat çalışma'],
    ['Volvo CE', 'volvoce.com', 'EC230 Electric, L25 Electric', '264 kWh', '5 saat çalışma'],
    ['Komatsu', 'komatsu.com', 'PC30E-5', '-', 'Mini ekskavatör'],
    ['Hitachi', 'hitachicm.com', 'ZX55U-6EB', '39.4 kWh', '2 saat, kablolu opsiyonu'],
    ['JCB', 'jcb.com', '19C-1E', '-', 'Mini ekskavatör'],
    ['Bobcat', 'bobcat.com', 'E10e, T7X', '-', '8 saat çalışma'],
    ['SANY', 'sanygroup.com', 'SY19E', '-', '6 saat, hızlı şarj'],
    ['Avant', 'avanttecno.com', 'E5, E6', '-', 'Kompakt loader'],
    ['HeavyEquipmentData (Kaynak)', 'heavyequipmentdata.com', 'API/Excel', '-', '60+ yıl veri'],
    ['SPEC CHECK (Kaynak)', 'speccheck.com', 'API/Portal', '-', 'Rekabet analizi'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws10.cell(row=i, column=j, value=val)
style_data(ws10, 2, len(data)+1, 5)
set_column_widths(ws10, [25, 30, 30, 12, 25])

# ===================== SAYFA 11: ATV/UTV/GOLF =====================
ws11 = wb.create_sheet("ATV UTV Golf")
headers = ['Üretici', 'URL', 'Modeller', 'Menzil', 'Özellikler']
for i, h in enumerate(headers, 1):
    ws11.cell(row=1, column=i, value=h)
style_header(ws11, 1, 5)

data = [
    ['Polaris', 'polaris.com/off-road/electric-vehicles', 'Ranger Kinetic', '80 mil', 'Zero Motorcycles ortaklığı'],
    ['Club Car', 'clubcar.com', 'Carryall, Onward', '-', 'Lithium opsiyon'],
    ['E-Z-GO', 'ezgo.com', 'RXV, TXT', '-', 'Golf cart lideri'],
    ['Yamaha', 'yamahamotorsports.com', 'Electric golf carts', '-', 'Golf cart'],
    ['HuntVe', 'huntve.com', '4x4 elektrikli UTV', '-', '10 yıl batarya garantisi'],
    ['Greenworks Commercial', 'greenworkscommercial.com', 'Utility vehicles', '-', 'Ticari kullanım'],
    ['John Deere', 'deere.com', 'Gator Electric', '-', 'Tarım/utility'],
    ['BIS Research (Kaynak)', 'bisresearch.com', 'Pazar analizi', '-', '$6.81B 2032 projeksiyonu'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws11.cell(row=i, column=j, value=val)
style_data(ws11, 2, len(data)+1, 5)
set_column_widths(ws11, [22, 45, 25, 10, 30])

# ===================== SAYFA 12: TÜRKİYE ÜRETİM/SATIŞ =====================
ws12 = wb.create_sheet("TÜRKİYE")
ws12['A1'] = "TÜRKİYE ELEKTRİKLİ ARAÇ ÜRETİM VE SATIŞ VERİLERİ"
ws12['A1'].font = Font(bold=True, size=14, color='C00000')
ws12.merge_cells('A1:H1')

# Üretim tablosu
ws12['A3'] = "TÜRKİYE'DE ÜRETİLEN ELEKTRİKLİ ARAÇLAR"
ws12['A3'].font = Font(bold=True, size=12)
headers = ['Marka', 'Model', 'Tip', 'Fabrika', 'Şehir', 'Yerlilik %', '2024 Satış', '2025 Satış']
for i, h in enumerate(headers, 1):
    ws12.cell(row=4, column=i, value=h)
style_header(ws12, 4, 8, turkey_fill)

data = [
    ['TOGG', 'T10X', 'C-SUV', 'TOGG Gemlik', 'Bursa', '72%', '30,093', '22,131 (9 ay)'],
    ['TOGG', 'T10F', 'Fastback Sedan', 'TOGG Gemlik', 'Bursa', '72%', '-', '2025 başladı'],
    ['Ford Otosan', 'E-Transit', 'Van', 'Ford Otosan', 'Kocaeli', '65%+', '-', 'Üretimde'],
    ['Ford Otosan', 'E-Tourneo', 'Minibüs', 'Ford Otosan', 'Kocaeli', '65%+', '-', 'Planlanan'],
    ['Karsan', 'e-Jest', 'Minibüs', 'Karsan', 'Bursa', '60%+', '-', 'İhracat'],
    ['Karsan', 'e-ATAK', 'Midibüs', 'Karsan', 'Bursa', '60%+', '-', 'İhracat'],
    ['Karsan', 'e-ATA', 'Otobüs', 'Karsan', 'Bursa', '60%+', '-', 'İhracat'],
    ['Otokar', 'e-Kent', 'Otobüs', 'Otokar', 'Sakarya', '-', '-', 'Üretimde'],
    ['BMC', 'Procity Electric', 'Otobüs', 'BMC', 'İzmir', '-', '-', 'Üretimde'],
]

for i, row in enumerate(data, 5):
    for j, val in enumerate(row, 1):
        ws12.cell(row=i, column=j, value=val)
style_data(ws12, 5, 13, 8)

# Satış tablosu
ws12['A16'] = "TÜRKİYE'DE EN ÇOK SATAN ELEKTRİKLİ ARAÇLAR (2025 Ocak-Eylül)"
ws12['A16'].font = Font(bold=True, size=12)
headers2 = ['Sıra', 'Marka', 'Model', 'Satış Adedi', 'Menşe', 'Üretim Yeri']
for i, h in enumerate(headers2, 1):
    ws12.cell(row=17, column=i, value=h)
style_header(ws12, 17, 6)

data2 = [
    ['1', 'Tesla', 'Model Y', '27,420', 'ABD', 'Çin/ABD'],
    ['2', 'TOGG', 'T10X', '22,131', 'TÜRKİYE', 'TÜRKİYE 🇹🇷'],
    ['3', 'MINI', 'Countryman', '7,028', 'İngiltere', 'Almanya'],
    ['4', 'Kia', 'EV3', '5,945', 'G.Kore', 'G.Kore'],
    ['5', 'BMW', 'iX1', '5,500+', 'Almanya', 'Almanya'],
    ['6', 'Volvo', 'EX30', '4,500+', 'İsveç', 'Çin'],
    ['7', 'BYD', 'Atto 3', '4,000+', 'Çin', 'Çin'],
    ['8', 'Mercedes', 'EQA', '3,500+', 'Almanya', 'Almanya'],
]

for i, row in enumerate(data2, 18):
    for j, val in enumerate(row, 1):
        cell = ws12.cell(row=i, column=j, value=val)
        if 'TÜRKİYE' in str(val):
            cell.font = Font(bold=True, color='C00000')
style_data(ws12, 18, 25, 6)

# Pazar istatistikleri
ws12['A28'] = "TÜRKİYE EV PAZAR İSTATİSTİKLERİ"
ws12['A28'].font = Font(bold=True, size=12)
stats = [
    ['Metrik', '2024', '2025 (9 ay)', 'Değişim'],
    ['Toplam EV Satışı', '105,315', '133,781', '+27%'],
    ['EV Pazar Payı', '~12%', '18%', '+6 puan'],
    ['Trafikteki EV Sayısı', '~200,000', '321,710', '+60%'],
    ['Şarj Soketi (AC)', '~15,000', '18,888', '+26%'],
    ['Şarj Soketi (DC)', '~10,000', '13,794', '+38%'],
    ['Toplam Şarj Noktası', '~25,000', '32,682', '+31%'],
]
for i, row in enumerate(stats, 29):
    for j, val in enumerate(row, 1):
        cell = ws12.cell(row=i, column=j, value=val)
        if i == 29:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = header_font

style_data(ws12, 30, 35, 4)
set_column_widths(ws12, [18, 20, 18, 18, 15, 15, 15, 18])

# ===================== SAYFA 13: ŞARJ VERİSİ =====================
ws13 = wb.create_sheet("Şarj Verileri")
headers = ['Kaynak', 'URL', 'Kapsam', 'Kayıt Sayısı', 'Format', 'Lisans']
for i, h in enumerate(headers, 1):
    ws13.cell(row=1, column=i, value=h)
style_header(ws13, 1, 6)

data = [
    ['Figshare - China EV', 'figshare.com/28182251', 'Çin, 13 istasyon, 2 yıl', '441,077 işlem', 'CSV/JSON', 'CC BY 4.0'],
    ['Figshare - Korea EV', 'figshare.com/22495141', 'G.Kore, 2,238 kullanıcı', '72,856 oturum', 'CSV', 'CC BY 4.0'],
    ['ST-EVCDP (GitHub)', 'github.com/IntelligentSystemsLab/ST-EVCDP', 'Shenzhen, 30 gün', '18,061 şarj noktası', 'CSV', 'Açık'],
    ['UrbanEV (GitHub)', 'github.com/IntelligentSystemsLab/UrbanEV', 'Shenzhen, 6 ay', '24,798 şarj noktası', 'CSV', 'Açık'],
    ['ACN-Data (Caltech)', 'ev.caltech.edu/dataset', 'Caltech şarj ağı', 'Yıllar', 'API', 'Akademik'],
    ['IEEE Dataport', 'ieee-dataport.org', 'Çeşitli', 'Çeşitli', 'Çeşitli', 'Akademik'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws13.cell(row=i, column=j, value=val)
style_data(ws13, 2, len(data)+1, 6)

ws13['A10'] = "TÜRKİYE YERLİ ŞARJ İSTASYONU ÜRETİCİLERİ"
ws13['A10'].font = Font(bold=True, size=12, color='C00000')
headers2 = ['Üretici', 'Marka', 'Kapasite', 'Özellikler']
for i, h in enumerate(headers2, 1):
    ws13.cell(row=11, column=i, value=h)
style_header(ws13, 11, 4, turkey_fill)

data2 = [
    ['Vestel/ZES', 'ZES', '180 kW', 'AC+DC, E.ON/Iberdrola ihracat'],
    ['Gersan Elektrik', 'G-Charge', '50-350 kW', '92 istasyon, %70 yerlilik'],
    ['Aspower', 'Aspower', '7-350 kW', 'Yerli üretim'],
    ['Aselsan', 'Aselsan', 'AC/DC', 'Savunma sanayi'],
]
for i, row in enumerate(data2, 12):
    for j, val in enumerate(row, 1):
        ws13.cell(row=i, column=j, value=val)
style_data(ws13, 12, 15, 4)
set_column_widths(ws13, [25, 45, 20, 15, 15, 15])

# ===================== SAYFA 14: VERİ KAYNAKLARI ÖZET =====================
ws14 = wb.create_sheet("Veri Kaynakları Özet")
headers = ['Kategori', 'Öncelikli Kaynak', 'Format', 'Erişim', 'Tahmini Süre', 'Maliyet']
for i, h in enumerate(headers, 1):
    ws14.cell(row=1, column=i, value=h)
style_header(ws14, 1, 6)

data = [
    ['Binek', 'Open EV Data + EVKX', 'JSON/MD', 'Ücretsiz', '1 gün', '$0'],
    ['Motosiklet', 'TeoAlida', 'Excel/CSV', 'Ücretli', '1 gün', '$50-200'],
    ['E-Scooter', 'Üretici siteleri scrape', 'Web', 'Ücretsiz', '1 hafta', '$0'],
    ['E-Bisiklet', 'Bikes.fan', 'Web', 'Ücretsiz', '1 hafta', '$0'],
    ['E-Otobüs', 'Sustainable Bus + BYD', 'Web', 'Ücretsiz', '3 gün', '$0'],
    ['E-Kamyon', 'Üretici siteleri', 'Web/PDF', 'Ücretsiz', '3 gün', '$0'],
    ['E-Tekne', 'Plugboats', 'Web', 'Ücretsiz', '2 gün', '$0'],
    ['eVTOL', 'eVTOL.news', 'Web', 'Ücretsiz', '1 gün', '$0'],
    ['İş Makinesi', 'HeavyEquipmentData', 'Excel/API', 'Ücretli', '1 gün', '$100-500'],
    ['ATV/UTV', 'Üretici siteleri', 'Web', 'Ücretsiz', '2 gün', '$0'],
    ['Türkiye', 'ODMD + Üretici siteleri', 'Web/PDF', 'Ücretsiz', '1 gün', '$0'],
    ['Şarj Verisi', 'Figshare + GitHub', 'CSV', 'Ücretsiz', '1 gün', '$0'],
]

for i, row in enumerate(data, 2):
    for j, val in enumerate(row, 1):
        ws14.cell(row=i, column=j, value=val)
style_data(ws14, 2, len(data)+1, 6)
set_column_widths(ws14, [15, 30, 12, 12, 15, 12])

# Kaydet
wb.save('/home/claude/elektrikli_arac_veritabanlari.xlsx')
print("Excel dosyası oluşturuldu: elektrikli_arac_veritabanlari.xlsx")